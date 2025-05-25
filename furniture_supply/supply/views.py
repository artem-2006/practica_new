from django.shortcuts import render, redirect, get_object_or_404
from .models import SupplyOrder, Product, Supplier
from django import forms
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter


class SupplyOrderForm(forms.ModelForm):
    class Meta:
        model = SupplyOrder
        fields = ['product', 'supplier', 'quantity', 'status']

    def __init__(self, *args, **kwargs):
        super(SupplyOrderForm, self).__init__(*args, **kwargs)
        for field in self.fields.values():
            field.widget.attrs.update({'class': 'form-control rounded-3 py-2'})
        if 'product' in self.fields:
            self.fields['product'].empty_label = "Выберите товар"
        if 'supplier' in self.fields:
            self.fields['supplier'].empty_label = "Выберите поставщика"


def admin_required(view_func):
    return user_passes_test(lambda u: u.is_active and u.is_staff)(view_func)


@admin_required
def order_list(request):
    orders = SupplyOrder.objects.all().order_by('-order_date')
    return render(request, 'supply/order_list.html', {'orders': orders})

@admin_required
def order_create(request):
    if request.method == 'POST':
        form = SupplyOrderForm(request.POST)
        if form.is_valid():
            supply_order = form.save(commit=False)
            product = supply_order.product
            if supply_order.quantity > product.stock_quantity:
                form.add_error('quantity', 'Нельзя заказать больше товара, чем есть на складе.')
            else:
                supply_order.save()
                
                product.stock_quantity -= supply_order.quantity
                product.save()
                return redirect('order_list')
    else:
        form = SupplyOrderForm()
    return render(request, 'supply/order_form.html', {'form': form})


@admin_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'supply/product_list.html', {'products': products})

@admin_required
def product_create(request):
    class ProductForm(forms.ModelForm):
        class Meta:
            model = Product
            fields = ['name', 'description', 'price', 'stock_quantity']

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            for field in self.fields.values():
                field.widget.attrs.update({'class': 'form-control rounded-3 py-2'})

    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'supply/product_form.html', {'form': form})


@admin_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'supply/supplier_list.html', {'suppliers': suppliers})

@admin_required
def supplier_create(request):
    class SupplierForm(forms.ModelForm):
        class Meta:
            model = Supplier
            fields = ['name', 'contact_info']

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            for field in self.fields.values():
                field.widget.attrs.update({'class': 'form-control rounded-3 py-2'})

    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'supply/supplier_form.html', {'form': form})


@admin_required
def order_report(request):
    orders = SupplyOrder.objects.all().order_by('-order_date')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы поставок"

    headers = ['Товар', 'Поставщик', 'Количество', 'Статус', 'Дата заказа']
    ws.append(headers)

    for order in orders:
        ws.append([
            order.product.name,
            order.supplier.name,
            order.quantity,
            order.get_status_display(),
            order.order_date.strftime('%Y-%m-%d %H:%M:%S') if order.order_date else ''
        ])

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=orders_report.xlsx'
    wb.save(response)
    return response


@admin_required
def product_report(request):
    products = Product.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    headers = ['Название', 'Описание', 'Цена', 'Количество на складе']
    ws.append(headers)

    for product in products:
        ws.append([
            product.name,
            product.description,
            product.price,
            product.stock_quantity
        ])

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products_report.xlsx'
    wb.save(response)
    return response


@admin_required
def supplier_report(request):
    suppliers = Supplier.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Поставщики"

    headers = ['Название', 'Контактная информация']
    ws.append(headers)

    for supplier in suppliers:
        ws.append([
            supplier.name,
            supplier.contact_info
        ])

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 30

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=suppliers_report.xlsx'
    wb.save(response)
    return response
